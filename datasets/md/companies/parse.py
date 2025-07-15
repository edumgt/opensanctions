import re
from datetime import datetime, timedelta
from lxml import html
from openpyxl import Workbook, load_workbook
from rigour.mime.types import XLSX
from typing import Optional, Dict, Any, List
from urllib.parse import urljoin

from zavod import Context, Entity, helpers as h

# This parser handles two separate data sources:
# 1. State Register of legal entities in the Republic of Moldova
# 2. Moldovan Registry of Non-Profit Organizations
# Although the extraction logic for each is distinct and unrelated,
# they are grouped under the same parser to allow unified processing
# and enrichment through a shared enricher.

IGNORE_COLUMNS = [
    "Codul unităţii administrativ-teritoriale",
    "Genuri de activitate nelicentiate",
    "Genuri de activitate licentiate",
]
NONPROFITS_URL = "https://dataset.gov.md/dataset/18516-date-din-registrul-de-stat-al-unitatilor-de-drept-privind-organizatiile-necomerciale"


def read_ckan(context: Context) -> str:
    if context.dataset.url is None:
        raise RuntimeError("No dataset url")
    path = context.fetch_resource("dataset.html", context.dataset.url)
    with open(path, "r") as fh:
        doc = html.fromstring(fh.read())

    resource_url = None
    for res_anchor in doc.findall('.//li[@class="resource-item"]/a'):
        res_href = res_anchor.get("href", "")
        resource_url = urljoin(context.dataset.url, res_href)

    if resource_url is None:
        raise RuntimeError("No resource URL on data catalog page!")

    path = context.fetch_resource("resource.html", resource_url)
    with open(path, "r") as fh:
        doc = html.fromstring(fh.read())

    for action_anchor in doc.findall('.//div[@class="actions"]//a'):
        return action_anchor.get("href")

    raise RuntimeError("No data URL on data resource page!")


def parse_directors(
    context: Context, company: Entity, directors: Optional[str]
) -> None:
    if directors is None:
        return
    for director in directors.split("],"):
        # if "[" not in director:
        #     print(director, directors)
        #     continue
        role = None
        try:
            director, role = director.rsplit("[", 1)
            role = role.replace("]", "").strip()
        except ValueError:
            pass

        director = director.strip()
        if len(director) < 3:
            continue

        dir = context.make("LegalEntity")
        dir.id = context.make_id(company.id, director)
        dir.add("name", director)
        context.emit(dir)

        dship = context.make("Directorship")
        dship.id = context.make_id("Directorship", company.id, director, role)
        dship.add("organization", company.id)
        dship.add("director", dir.id)
        dship.add("role", role)
        context.emit(dship)


def parse_founders(context: Context, company: Entity, founders: Optional[str]) -> None:
    if founders is None:
        return
    if isinstance(founders, int):
        context.log.info("last line: %r" % founders)
        return
    for founder in founders.split("),"):
        founder = founder.replace(")", "")
        percentage = None
        if "(" in founder:
            founder, percentage = founder.rsplit("(", 1)

        founder = founder.strip()
        found = context.make("LegalEntity")
        found.id = context.make_id(company.id, founder)
        found.add("name", founder)
        if not found.has("name"):
            continue
        context.emit(found)

        own = context.make("Ownership")
        own.id = context.make_id("Ownership", company.id, founder)
        own.add("asset", company.id)
        own.add("owner", found.id)
        own.add("role", percentage)
        context.emit(own)


def parse_owners(context: Context, company: Entity, owners: Optional[str]) -> None:
    if owners is None:
        return
    for owner in owners.split("),"):
        owner = owner.replace(")", "")
        country = None
        if "(" in owner:
            owner, country = owner.rsplit("(", 1)

        owner = owner.strip()
        bo = context.make("LegalEntity")
        bo.id = context.make_id(company.id, owner)
        bo.add("name", owner)
        bo.add("country", country)
        if country is not None and not bo.has("country"):
            context.log.warn("Unknown country code", country=country)
        if not bo.has("name"):
            continue
        context.emit(bo)

        own = context.make("Ownership")
        own.id = context.make_id("Ownership", company.id, owner)
        own.add("asset", company.id)
        own.add("owner", bo.id)
        own.add("role", "beneficiarilor efectivi")
        context.emit(own)


def parse_company(context: Context, data: Dict[str, Any]) -> None:
    idno = data.pop("IDNO/ Cod fiscal")
    name = data.pop("Denumirea completă")
    address = data.pop("Adresa")
    company = context.make("Company")
    if idno is not None:
        company.id = f"oc-companies-md-{idno}"
    else:
        company.id = context.make_id(name, address)
    if company.id is None:
        context.log.error(
            "Cannot generate key",
            idno=idno,
            name=name,
            address=address,
        )
        return
    company.add("name", name)
    company.add("taxNumber", idno)
    company.add("incorporationDate", data.pop("Data înregistrării"))
    company.add("dissolutionDate", data.pop("Data lichidării"))
    company.add("jurisdiction", "md")
    company.add("address", address)
    company.add("legalForm", data.pop("Forma org./jurid."))
    parse_directors(context, company, data.pop("Lista conducătorilor"))
    parse_founders(context, company, data.pop("Lista fondatorilor"))
    parse_owners(context, company, data.pop("Lista beneficiarilor efectivi", None))

    context.emit(company)
    context.audit_data(data, ignore=IGNORE_COLUMNS)


def parse_companies(context: Context, book: Workbook) -> None:
    headers: Optional[List[str]] = None
    for idx, row in enumerate(book["Company"].iter_rows()):
        cells = [c.value for c in row]
        if headers is None:
            if "Denumirea completă" in cells:
                headers = []
                for cell in cells:
                    header = str(cell).split(" (", 1)[0]
                    headers.append(header)
            continue
        data = dict(zip(headers, cells))
        parse_company(context, data)
        if idx > 0 and idx % 10000 == 0:
            context.log.info("Read %d companies..." % idx)


def get_most_recent_link(context, doc):
    # Select all .xlsx resource links within resource-item elements
    links = doc.xpath(
        "//li[contains(@class, 'resource-item')]//a[contains(@class, 'resource-url-analytics') and contains(@href, '.xlsx')]"
    )
    dated_links = []

    for link in links:
        href = link.get("href")
        # The XPath selects all downloadable XLSX links on the page, including outdated ones.
        # However, we're only interested in the most recent file, so we filter out known outdated links,
        # such as one with a hardcoded 2024 date and any from 2019.
        # These older links follow a different URL format and lack consistent date patterns.
        # Since we assert that the latest link must be within the last 30 days,
        # it's safe and more efficient to exclude them upfront to simplify parsing.
        if "17.06.2024" in href or re.search(r"resources/2019-\d{2}", href):
            continue
        # Full date: YYYY.MM.DD
        match = re.search(r"(\d{4})[.](\d{2})[.](\d{2})", href)
        if match:
            date_obj = datetime.strptime(match.group(0), "%Y.%m.%d")
            dated_links.append((date_obj, href))
        else:
            context.log.warning(f"Link {href} does not contain a date")
    # Select the most recent dated link
    latest_link = max(dated_links, key=lambda x: x[0])
    # Ensure the latest file is from the last 30 days
    assert datetime.now() - timedelta(days=30) < latest_link[0]

    return latest_link[1]


def crawl_row(context, row):
    tax_number = row.pop("tax_number")
    name = row.pop("name")
    director = row.pop("director")
    entity = context.make("Organization")
    entity.id = context.make_id(tax_number, name)
    entity.add("name", name)
    entity.add("legalForm", row.pop("legal_form"))
    entity.add("country", "md")
    entity.add("taxNumber", tax_number)
    address = h.make_address(
        context,
        full=row.pop("address"),
        place=row.pop("admin_unit_code"),
    )
    h.copy_address(entity, address)
    h.apply_date(entity, "incorporationDate", row.pop("incorporation_date"))
    h.apply_date(entity, "dissolutionDate", row.pop("dissolution_date"))

    if director:
        dir = context.make("Person")
        dir.id = context.make_id(director)
        dir.add("name", director)

        directorship = context.make("Directorship")
        directorship.id = context.make_id(entity.id, dir.id)
        directorship.add("organization", entity.id)
        directorship.add("director", dir.id)

        context.emit(dir)
        context.emit(directorship)

    context.emit(entity)
    context.audit_data(row)


def crawl(context: Context) -> None:
    # Companies
    data_url = read_ckan(context)
    data_path = context.fetch_resource("data.xlsx", data_url)
    # data_path = context.get_resource_path("data.xlsx")
    wb = load_workbook(data_path, read_only=True, data_only=True)
    parse_companies(context, wb)

    # Nonprofits
    doc = context.fetch_html(NONPROFITS_URL, cache_days=1)
    data_url = get_most_recent_link(context, doc)
    path = context.fetch_resource("list.xlsx", data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)
    wb = load_workbook(path, read_only=True)
    assert set(wb.sheetnames) == {wb.active.title}
    for row in h.parse_xlsx_sheet(
        context, wb["organizations"], skiprows=4, header_lookup="columns"
    ):
        crawl_row(context, row)
